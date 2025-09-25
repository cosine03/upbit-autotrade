# -*- coding: utf-8 -*-
import os
os.environ.setdefault("MPLBACKEND", "Agg")  # non-interactive backend

import argparse, os as _os
import pandas as pd
import numpy as np

# ---- helpers ----
def compute_equity_and_mdd(trades: pd.DataFrame, group_cols=("strategy","expiry_h")):
    """
    trades columns expected:
      - ts_entry (optional), event, expiry_h, net (per trade PnL), strategy
    Return dict: {(strategy, expiry): (equity_df, mdd_float)}
    """
    out = {}
    if "net" not in trades.columns:
        return out
    # sort by time if available
    if "ts_entry" in trades.columns:
        trades["ts_entry"] = pd.to_datetime(trades["ts_entry"], errors="coerce", utc=True)
        trades = trades.sort_values(["strategy","expiry_h","ts_entry"])
    for keys, g in trades.groupby(list(group_cols)):
        eq = g["net"].cumsum().reset_index(drop=True)
        peak = eq.cummax()
        dd = eq - peak
        mdd = dd.min() if len(dd) else 0.0
        equity = pd.DataFrame({"step": np.arange(len(eq)), "equity": eq.values})
        out[keys] = (equity, float(mdd))
    return out

def save_bar(ax, df, title, ycol):
    for col in df.columns[1:]:
        ax.bar([f"{x} {col}" for x in df[df.columns[0]]], df[col])
    ax.set_title(title); ax.set_xlabel(f"{df.columns[0]} x strategy"); ax.set_ylabel(ycol)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in1", required=True, help="summary csv for breakout-only")
    ap.add_argument("--in2", required=True, help="summary csv for boxin-linebreak subset")
    ap.add_argument("--trades1", required=True, help="trades csv for breakout-only")
    ap.add_argument("--trades2", required=True, help="trades csv for boxin-linebreak")
    ap.add_argument("--out", required=True, help="output xlsx")
    ap.add_argument("--tag", default="", help="date tag to annotate")
    args = ap.parse_args()

    # summaries
    b = pd.read_csv(args.in1)
    s = pd.read_csv(args.in2)
    keep = ["event","expiry_h","trades","win_rate","avg_net","median_net","total_net"]
    b = b[keep].copy(); s = s[keep].copy()
    b["strategy"] = "breakout_only"; s["strategy"] = "boxin_linebreak"
    both = pd.concat([b,s], ignore_index=True)

    # trades
    tb = pd.read_csv(args.trades1)
    ts = pd.read_csv(args.trades2)
    tb["strategy"] = "breakout_only"
    ts["strategy"] = "boxin_linebreak"
    trades = pd.concat([tb, ts], ignore_index=True)

    # write sheets
    with pd.ExcelWriter(args.out, engine="openpyxl") as writer:
        b.to_excel(writer, sheet_name="breakout_only", index=False)
        s.to_excel(writer, sheet_name="boxin_linebreak", index=False)
        both.to_excel(writer, sheet_name="combined", index=False)
        trades.to_excel(writer, sheet_name="trades", index=False)

        # pivot summaries
        pivot_avg = both.pivot_table(index=["expiry_h"], columns="strategy", values="avg_net", aggfunc="mean").reset_index()
        pivot_tot = both.pivot_table(index=["expiry_h"], columns="strategy", values="total_net", aggfunc="sum").reset_index()
        pivot_wr  = both.pivot_table(index=["expiry_h"], columns="strategy", values="win_rate", aggfunc="mean").reset_index()
        pivot_avg.to_excel(writer, sheet_name="pivots", startrow=0, index=False)
        pivot_tot.to_excel(writer, sheet_name="pivots", startrow= pivot_avg.shape[0]+3, index=False)
        pivot_wr.to_excel(writer,  sheet_name="pivots", startrow= pivot_avg.shape[0]+pivot_tot.shape[0]+6, index=False)

    # charts
    import matplotlib.pyplot as plt
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage

    base = _os.path.splitext(args.out)[0]
    png1, png2, png3 = base+"_avg_net.png", base+"_total_net.png", base+"_win_rate.png"

    fig, ax = plt.subplots(figsize=(8,5)); save_bar(ax, pivot_avg, f"AVG_NET by expiry ({args.tag})", "avg_net"); plt.tight_layout(); plt.savefig(png1); plt.close()
    fig, ax = plt.subplots(figsize=(8,5)); save_bar(ax, pivot_tot, f"TOTAL_NET by expiry ({args.tag})", "total_net"); plt.tight_layout(); plt.savefig(png2); plt.close()
    fig, ax = plt.subplots(figsize=(8,5)); save_bar(ax, pivot_wr,  f"WIN_RATE by expiry ({args.tag})", "win_rate"); plt.tight_layout(); plt.savefig(png3); plt.close()

    # equity & MDD
    emap = compute_equity_and_mdd(trades)
    # 저장용 이미지들
    eq_imgs = []
    for (strategy, expiry), (eqdf, mdd) in emap.items():
        fn = f"{base}_equity_{strategy}_{expiry}.png"
        plt.figure(figsize=(8,4))
        plt.plot(eqdf["step"], eqdf["equity"])
        plt.title(f"Equity Curve - {strategy} - {expiry}h | MDD={mdd:.4f}")
        plt.xlabel("trade idx"); plt.ylabel("cumulative net")
        plt.tight_layout(); plt.savefig(fn); plt.close()
        eq_imgs.append((fn, strategy, expiry, mdd))

    # insert images
    wb = load_workbook(args.out)
    ws = wb.create_sheet("charts")
    ws["A1"] = f"Tag: {args.tag}"
    row = 3
    for p in (png1, png2, png3):
        if _os.path.exists(p):
            img = XLImage(p); img.anchor = f"A{row}"; ws.add_image(img); row += 25

    # equity/mdd 섹션
    ws2 = wb.create_sheet("equity_mdd")
    ws2["A1"] = "strategy"; ws2["B1"] = "expiry_h"; ws2["C1"] = "MDD"
    r = 2
    for fn, strategy, expiry, mdd in eq_imgs:
        ws2[f"A{r}"] = strategy
        ws2[f"B{r}"] = float(expiry)
        ws2[f"C{r}"] = float(mdd)
        # 그림
        if _os.path.exists(fn):
            img = XLImage(fn); img.anchor = f"E{r}"; ws2.add_image(img)
        r += 20

    wb.save(args.out)
    print("[REPORT] saved ->", args.out)

if __name__ == "__main__":
    main()